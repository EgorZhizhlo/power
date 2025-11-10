import io
from math import ceil
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, NamedStyle


THIN = Side(border_style="thin",   color="000000")
MEDIUM = Side(border_style="medium", color="000000")
ROW_HEIGHT_PT = 15
COL_B_W, COL_C_W = 52, 22
COL_B_CH = int(COL_B_W / 0.9)
COL_C_CH = int(COL_C_W / 0.9)

hdr = NamedStyle("hdr")
hdr.number_format = "@"
hdr.font = Font(bold=True)
hdr.alignment = Alignment(
    wrap_text=True, horizontal="center", vertical="center")

dat = NamedStyle("dat")
dat.number_format = "@"
dat.alignment = Alignment(wrap_text=True, vertical="top")

fac = NamedStyle("fac")
fac.number_format = "@"
fac.alignment = Alignment(wrap_text=True, vertical="top")


def register_styles(wb: Workbook):
    for style in (hdr, dat, fac):
        if style.name not in wb.named_styles:
            wb.add_named_style(style)


def make_border(l=THIN, r=THIN, t=THIN, b=THIN) -> Border:
    return Border(left=l, right=r, top=t, bottom=b)


def wrapped_lines(text: str, width: int) -> int:
    return sum(max(1, ceil(len(line) / width)) for line in text.splitlines() or [""])


def prepare_orders(input_data: list[dict]) -> list[dict]:
    orders = []
    for item in input_data:
        phone = item.get("phone_number", "")
        sec_phone_number = item.get("sec_phone_number", "")
        address = ", ".join(
            filter(None, [item.get("city_name", ""), item.get("address", "")]))
        cnt = item.get("counter_number", 0)
        wt = item.get("water_type", "")
        price = item.get("price", 0.0)
        additional_info = item.get("additional_info") or ""
        comment = f"{cnt}-{wt}; {int(price)}₽; {additional_info}"
        orders.append({
            "phone": phone, "sec_phone": sec_phone_number,
            "address": address, "comment": comment})
    return orders


def build_header(ws, report_date, route_name, employee_full_name):
    text = (
        f"Путевой (заявочный) лист на {report_date:%d.%m.%Y}; "
        f"МАРШРУТ {route_name}; {employee_full_name}"
    )

    ws.merge_cells("A1:C1")
    c = ws["A1"]
    c.value = text
    c.font = Font(bold=True)
    c.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

    # считаем сколько строк займёт текст при ширине A+B+C
    total_width = COL_B_CH + COL_C_CH + int(9 / 0.9)  # ширина A+B+C в символах
    lines = wrapped_lines(text, total_width)
    ws.row_dimensions[1].height = ROW_HEIGHT_PT * lines

    for col, txt in (
        ("A", "№ п/п"),
        ("B", "(Тел./Адрес/Кол-во/Сумма) в заявке"),
        ("C", "(Кол-во/Сумма) Факт")
    ):
        ws.merge_cells(f"{col}3:{col}4")
        c = ws[f"{col}3"]
        c.value = txt
        c.style = hdr


def build_table(ws, orders, start_row=5):
    for idx, o in enumerate(orders, 1):
        r = start_row + idx - 1
        ws.cell(r, 1, idx).style = dat

        b_list = list(
            filter(
                None, [o["phone"], o["sec_phone"], o["address"], o["comment"]]
            )
        )
        b_text = "; ".join(b_list)

        ws.cell(r, 2, b_text).style = dat
        ws.cell(r, 3, "Кол-во____\nСумма____").style = fac

        # подгоняем высоту
        lines_b = wrapped_lines(b_text, COL_B_CH)
        lines_c = wrapped_lines("Кол-во____\nСумма____", COL_C_CH)
        ws.row_dimensions[r].height = ROW_HEIGHT_PT * max(lines_b, lines_c)


def build_footer(ws, footer_row):
    ws[f"A{footer_row}"] = "ИТОГО:"
    ws[f"A{footer_row}"].style = hdr

    foot_text = (
        "Кол-во поверенных СИ: ___     Кол-во непригодных СИ: ___\n"
        "Кол-во выданных актов: ___    Кол-во выданных актов ГВК: ___"
    )
    ws.merge_cells(f"B{footer_row}:C{footer_row}")
    ws[f"B{footer_row}"] = foot_text
    ws[f"B{footer_row}"].style = fac

    lines_f = wrapped_lines(foot_text, COL_B_CH + COL_C_CH)
    ws.row_dimensions[footer_row].height = ROW_HEIGHT_PT * lines_f + 4


def apply_borders(ws, max_row):
    for r in range(1, max_row + 1):
        for c in (1, 2, 3):
            ws.cell(r, c).border = make_border(
                l=MEDIUM if c == 1 else THIN,
                r=MEDIUM if c == 3 else THIN,
                t=MEDIUM if r == 1 else THIN,
                b=MEDIUM if r == max_row else THIN
            )


def set_sheet_options(ws, max_row):
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = COL_B_W
    ws.column_dimensions["C"].width = COL_C_W
    ws.freeze_panes = "A5"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_area = f"A1:C{max_row}"


def create_report_route_orders_list(
        input_data: list[dict],
        add_input_data: dict
) -> io.BytesIO:
    report_date = add_input_data["date"]
    route_name = add_input_data["route_name"]
    employee_full_name = add_input_data["employee_full_name"]
    route_additional_info = (add_input_data.get(
        "route_additional_info") or "").strip()

    orders = prepare_orders(input_data)
    wb = Workbook()
    ws = wb.active
    ws.title = "Маршрут"

    register_styles(wb)
    build_header(ws, report_date, route_name, employee_full_name)

    START_ROW = 5
    build_table(ws, orders, start_row=START_ROW)

    # первая свободная строка после заявок
    current_row = START_ROW + len(orders)

    # --- ВСТАВКА БЛОКА «Доп. информация» ПОСЛЕ ЗАЯВОК ---
    if route_additional_info:
        # Текст в A:C
        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row, end_column=3)
        bc_cell = ws.cell(current_row, 1, route_additional_info)
        bc_cell.style = fac  # обтекание + верхнее выравнивание

        # высота строки по количеству «визуальных» строк в объединённых B+C
        lines_bc = wrapped_lines(route_additional_info, COL_B_CH + COL_C_CH)
        ws.row_dimensions[current_row].height = ROW_HEIGHT_PT * lines_bc
        current_row += 1
    else:
        # Текст в A:C
        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row, end_column=3)
        bc_cell = ws.cell(current_row, 1, "")
        bc_cell.style = fac
        lines_bc = wrapped_lines("", COL_B_CH + COL_C_CH)
        ws.row_dimensions[current_row].height = ROW_HEIGHT_PT * lines_bc
        current_row += 1

    footer_row = current_row

    build_footer(ws, footer_row)
    apply_borders(ws, footer_row)
    set_sheet_options(ws, footer_row)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
