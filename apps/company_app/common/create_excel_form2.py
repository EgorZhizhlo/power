import io
from math import ceil
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, NamedStyle

THIN = Side(border_style="thin", color="000000")
ROW_HEIGHT_PT = 15

hdr = NamedStyle("hdr")
hdr.font = Font(name="Times New Roman", bold=False, size=10)
hdr.alignment = Alignment(
    wrap_text=True, horizontal="center", vertical="center")

dat = NamedStyle("dat")
dat.font = Font(name="Times New Roman", size=9)
dat.alignment = Alignment(wrap_text=True, horizontal="left", vertical="center")


def register_styles(wb: Workbook):
    for style in (hdr, dat):
        if style.name not in wb.named_styles:
            wb.add_named_style(style)


def make_border():
    return Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def wrapped_lines(text: str, width: int) -> int:
    return sum(max(1, ceil(len(line) / width)) for line in text.splitlines() or [""])


COLUMNS = [
    ("index", "№", 3),
    ("measurement_type", "Виды измерений, тип (группа) средств измерений", 9),
    ("standards", "Эталоны единиц величин и (или) СИ, тип (марка), "
                  "регистрационный номер в ФИФ по обеспечению единства измерений (при наличии)", 13),
    ("manufacturer", "Изготовитель (страна, наименование организации, год выпуска)", 11),
    ("commission_year", "Год ввода в эксплуатацию, инвентарный номер или "
                        "другая уникальная идентификация", 12),
    ("range", "Диапазон измерений", 13),
    ("accuracy", "Погрешность и (или) неопределённость (класс, разряд)", 13),
    ("certificate", "Свидетельство об аттестации эталонов единиц величин "
                    "(номер, дата, срок действия), сведения о результатах поверки СИ "
                    "в ФИФ по обеспечению единства измерений "
                    "(номер, дата, срок действия) и (или) сертификат калибровки", 16),
    ("ownership", "Право собственности и/или иное законное основание, "
                  "предусматривающее право владения и пользования (реквизиты документов)", 11),
    ("location", "Место установки или хранения", 9),
    ("note", "Примечание", 11),
]


def build_header(ws):
    # объединения
    ws.merge_cells("A1:A2")
    ws.merge_cells("B1:B2")
    ws.merge_cells("C1:C2")
    ws.merge_cells("D1:D2")
    ws.merge_cells("E1:E2")
    ws.merge_cells("F1:G1")
    ws.merge_cells("H1:H2")
    ws.merge_cells("I1:I2")
    ws.merge_cells("J1:J2")
    ws.merge_cells("K1:K2")

    headers_lvl1 = {
        "A1": "№",
        "B1": "Виды измерений, тип (группа) средств измерений",
        "C1": "Эталоны единиц величин и (или) СИ, тип (марка), "
              "регистрационный номер в ФИФ по обеспечению единства измерений (при наличии)",
        "D1": "Изготовитель (страна, наименование организации, год выпуска)",
        "E1": "Год ввода в эксплуатацию, инвентарный номер или другая уникальная идентификация",
        "F1": "Метрологические характеристики СИ",
        "H1": "Свидетельство об аттестации эталонов единиц величин (номер, дата, срок действия), "
              "сведения о результатах поверки СИ в ФИФ по обеспечению единства измерений "
              "(номер, дата, срок действия) и (или) сертификат калибровки (дата, срок действия, № при наличии)",
        "I1": "Право собственности и/или иное законное основание, предусматривающее право владения и пользования "
              "(реквизиты подтверждающих документов)",
        "J1": "Место установки или хранения",
        "K1": "Примечание",
    }

    for cell, text in headers_lvl1.items():
        ws[cell].value = text
        ws[cell].style = hdr

    ws["F2"].value = "Диапазон измерений"
    ws["F2"].style = hdr
    ws["G2"].value = "Погрешность и (или) неопределённость (класс, разряд)"
    ws["G2"].style = hdr

    for col, (_, _, width) in enumerate(COLUMNS, 1):
        ws.column_dimensions[chr(64 + col)].width = width

    def calc_height(row_idx, base_height=13, max_lines=7):
        max_lines_row = 1
        for c in range(1, len(COLUMNS) + 1):
            text = str(ws.cell(row_idx, c).value or "")
            width = int(ws.column_dimensions[chr(64 + c)].width / 0.9)
            max_lines_row = max(max_lines_row, wrapped_lines(text, width))
        max_lines_row = min(max_lines_row, max_lines)
        return base_height * max_lines_row

    h1 = calc_height(1, base_height=14)
    ws.row_dimensions[2].height = h1 * 1.5
    ws.row_dimensions[1].height = h1 * 0.6


def build_section(ws, section_title: str, rows: list[dict], start_row: int) -> int:
    ws.merge_cells(start_row=start_row, start_column=1,
                   end_row=start_row, end_column=len(COLUMNS))
    ws.cell(start_row, 1, section_title).style = hdr

    width_total = sum(int(ws.column_dimensions[chr(
        64 + c)].width / 0.95) for c in range(1, len(COLUMNS)+1))
    lines = wrapped_lines(section_title, width_total)
    ws.row_dimensions[start_row].height = ROW_HEIGHT_PT * lines

    for idx, row in enumerate(rows, 1):
        r = start_row + idx
        for col, (key, _, _) in enumerate(COLUMNS, 1):
            value = idx if key == "index" else row.get(key, "")
            cell = ws.cell(r, col)

            if key == "note" and isinstance(value, str) and value.startswith("http"):
                cell.value = value
                cell.hyperlink = value
                cell.font = Font(name="Times New Roman",
                                 color="0000FF", underline="single", size=9)
                cell.alignment = Alignment(
                    wrap_text=True, horizontal="left", vertical="center")
            else:
                cell.value = value
                cell.style = dat

        max_lines = 1
        for col in range(1, len(COLUMNS)+1):
            text = str(ws.cell(r, col).value or "")
            width = int(ws.column_dimensions[chr(64 + col)].width / 0.95)
            max_lines = max(max_lines, wrapped_lines(text, width))
        ws.row_dimensions[r].height = ROW_HEIGHT_PT * max_lines

    return start_row + len(rows) + 1


def apply_borders(ws):
    for r in range(1, ws.max_row + 1):
        for c in range(1, len(COLUMNS) + 1):
            ws.cell(r, c).border = make_border()


def set_sheet_options(ws):
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.1
    ws.page_margins.right = 0.1
    ws.page_margins.top = 0.1
    ws.page_margins.bottom = 0.1
    ws.print_area = f"A1:{chr(64+len(COLUMNS))}{ws.max_row}"


def create_table_report(sections: list[dict]) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Форма 2"

    register_styles(wb)
    build_header(ws)

    current_row = 3
    for section in sections:
        current_row = build_section(
            ws, section["section_title"], section["rows"], current_row)

    apply_borders(ws)
    set_sheet_options(ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
